"""
Database Population Script for Purchases, Ordered Slips, Suppliers, and Expenses
Run this with: python populate_purchases_and_expenses.py
"""

import os
import django
import sys
from datetime import datetime, date, timedelta
from decimal import Decimal

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp_system.settings')
django.setup()

from purchases.models import SupplierCompany, OrderedSlip, Purchase
from accounts.models import Expense, CashTransaction
from products.models import Product


def clear_existing_data():
    """Clear all existing data from the database for these models"""
    print("\nClearing existing purchases, ordered slips, supplier companies, and expenses...")
    Purchase.objects.all().delete()
    OrderedSlip.objects.all().delete()
    SupplierCompany.objects.all().delete()
    Expense.objects.all().delete()
    # Delete related CashTransactions to prevent orphaned ledger records
    CashTransaction.objects.filter(source_type__in=['purchase', 'expense', 'expense_void']).delete()
    print("Existing data cleared!")


def populate_suppliers():
    """Populate all suppliers from the Suppliers sheet in AlNoor_ERP_Dataset.xlsx"""
    print("\nPopulating Suppliers...")
    import openpyxl
    wb = openpyxl.load_workbook('AlNoor_ERP_Dataset.xlsx', read_only=True)
    sheet = wb['Suppliers']
    count = 0
    
    # Rows start at 3 (header is row 2)
    for r in range(3, sheet.max_row + 1):
        name = sheet.cell(row=r, column=2).value
        if not name:
            continue
        category = sheet.cell(row=r, column=3).value
        email = sheet.cell(row=r, column=6).value
        phone = sheet.cell(row=r, column=5).value
        
        supplier, created = SupplierCompany.objects.get_or_create(
            name=str(name).strip(),
            defaults={
                'category': str(category).strip() if category else '',
                'email': str(email).strip() if email else '',
                'contact_number': str(phone).strip() if phone else '',
            }
        )
        if created:
            count += 1
            
    print(f"Created {count} SupplierCompanies.")
    wb.close()


def populate_purchases_and_ordered_slips():
    """Populate purchases and ordered slips from AlNoor_ERP_Dataset.xlsx"""
    print("\nPopulating Purchases and Ordered Slips...")
    import openpyxl
    from django.utils.dateparse import parse_date
    wb = openpyxl.load_workbook('AlNoor_ERP_Dataset.xlsx', read_only=True)
    sheet = wb['Purchase Orders']
    os_count = 0
    p_count = 0
    
    # Rows start at 3
    for r in range(3, sheet.max_row + 1):
        po_num = sheet.cell(row=r, column=1).value
        if not po_num:
            continue
        
        date_val = sheet.cell(row=r, column=2).value
        supplier_name = sheet.cell(row=r, column=4).value
        product_sku = sheet.cell(row=r, column=5).value
        qty_ordered = sheet.cell(row=r, column=8).value
        unit_cost = sheet.cell(row=r, column=9).value
        status_str = sheet.cell(row=r, column=11).value
        payment_status_str = sheet.cell(row=r, column=12).value
        
        if not product_sku or qty_ordered is None or unit_cost is None:
            continue
            
        product_sku = str(product_sku).strip()
        product = Product.objects.filter(sku=product_sku).first()
        if not product:
            print(f"  Warning: Product SKU '{product_sku}' not found. Skipping PO {po_num}.")
            continue
            
        # Parse date
        if isinstance(date_val, datetime):
            date_obj = date_val.date()
        elif isinstance(date_val, date):
            date_obj = date_val
        else:
            date_obj = parse_date(str(date_val)) or date(2024, 1, 1)
            
        qty_ordered = int(qty_ordered)
        unit_cost = Decimal(str(unit_cost))
        total_cost = qty_ordered * unit_cost
        
        # Map Statuses
        if status_str == 'Received':
            os_status = OrderedSlip.STATUS_COMPLETED
            qty_received = qty_ordered
            received_at = datetime.combine(date_obj, datetime.min.time())
        elif status_str == 'Partially Received':
            os_status = OrderedSlip.STATUS_PARTIAL
            qty_received = qty_ordered // 2
            received_at = datetime.combine(date_obj, datetime.min.time())
        else:
            os_status = OrderedSlip.STATUS_PENDING
            qty_received = 0
            received_at = None
            
        due_date = datetime.combine(date_obj + timedelta(days=2), datetime.min.time())
        
        # OrderedSlip creation skipped as requested
        
        # If status is Received, also create a Purchase record
        if status_str == 'Received':
            if payment_status_str == 'Paid':
                pay_status = 'PAID'
            elif payment_status_str == 'Unpaid':
                pay_status = 'UNPAID'
            else:
                pay_status = 'PARTIAL'
                
            Purchase.objects.create(
                product=product,
                company_name=str(supplier_name).strip(),
                quantity_purchased=qty_ordered,
                unit_cost=unit_cost,
                total_cost=total_cost,
                purchase_date=date_obj,
                payment_status=pay_status,
                notes=f"Imported from Excel: {po_num}"
            )
            p_count += 1
            
    print(f"Created {os_count} OrderedSlips and {p_count} Purchases.")
    wb.close()


def populate_expenses():
    """Populate expenses from all monthly workbooks"""
    print("\nPopulating Expenses...")
    import openpyxl
    from django.utils.dateparse import parse_date
    
    # List of workbooks to read from
    workbooks = [
        "AlNoor_Financial_Summary_January_2026.xlsx",
        "AlNoor_Financial_Summary_February_2026.xlsx",
        "AlNoor_Financial_Summary_March_2026.xlsx",
        "AlNoor_Financial_Summary_April_2026.xlsx",
        "AlNoor_Financial_Summary.xlsx", # May
        "AlNoor_Financial_Summary_June_2026.xlsx",
    ]
    
    count = 0
    for wb_name in workbooks:
        if not os.path.exists(wb_name):
            print(f"  Warning: {wb_name} does not exist. Skipping.")
            continue
            
        try:
            wb = openpyxl.load_workbook(wb_name, read_only=True)
            if 'Expense Tracker' not in wb.sheetnames:
                wb.close()
                continue
            sheet = wb['Expense Tracker']
            
            # Expense data starts at row 3 and ends at row 15
            for r in range(3, 16):
                date_val = sheet.cell(row=r, column=1).value
                category = sheet.cell(row=r, column=2).value
                description = sheet.cell(row=r, column=3).value
                vendor = sheet.cell(row=r, column=4).value
                status_val = sheet.cell(row=r, column=5).value
                tax_amount = sheet.cell(row=r, column=6).value
                amount = sheet.cell(row=r, column=7).value
                
                if not category or amount is None:
                    continue
                    
                # Parse date
                if isinstance(date_val, datetime):
                    date_obj = date_val.date()
                elif isinstance(date_val, date):
                    date_obj = date_val
                else:
                    date_obj = parse_date(str(date_val))
                    
                if not date_obj:
                    continue
                    
                voided = (str(status_val).upper() == 'VOIDED')
                
                # Create Expense
                Expense.objects.create(
                    date=date_obj,
                    category=str(category).strip().upper(),
                    description=str(description).strip() if description else '',
                    vendor=str(vendor).strip() if vendor else '',
                    voided=voided,
                    void_reason='Voided transaction' if voided else '',
                    tax_amount=Decimal(str(tax_amount)) if tax_amount is not None else Decimal('0.00'),
                    amount=Decimal(str(amount)),
                    payment_method='CASH'
                )
                count += 1
            wb.close()
        except Exception as e:
            print(f"  Error reading {wb_name}: {e}")
            
    print(f"Created {count} Expenses.")


def main():
    print("\n" + "="*60)
    print("FINANCIAL DATA IMPORT & SEEDING SCRIPT")
    print("="*60)
    try:
        clear_existing_data()
        populate_suppliers()
        populate_purchases_and_ordered_slips()
        populate_expenses()
        print("\nIMPORT COMPLETED SUCCESSFULLY!")
    except Exception as e:
        print(f"\nError during execution: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
