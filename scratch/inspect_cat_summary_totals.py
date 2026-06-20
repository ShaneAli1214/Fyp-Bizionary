import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Category Summary"]
r = 28
print("Row 28 detail:")
for c in range(1, 6):
    print(f"Col {c}: value={sheet.cell(row=r, column=c).value}, type={type(sheet.cell(row=r, column=c).value)}")
wb.close()
