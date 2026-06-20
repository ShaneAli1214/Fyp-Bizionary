import openpyxl

def verify_file(path):
    print(f"\nVerifying file: {path}")
    wb = openpyxl.load_workbook(path, data_only=False)
    
    # Check Sales Data
    ws_sales = wb["Sales Data"]
    print(f"  Sales Data max_row: {ws_sales.max_row} (expected 163)")
    bad_cats_sales = []
    for r in range(3, ws_sales.max_row + 1):
        cat = ws_sales.cell(row=r, column=3).value
        if cat in ["Construction & Hardware", "Automobiles & Accessories"]:
            bad_cats_sales.append((r, cat))
    print(f"  Sales Data bad categories found: {len(bad_cats_sales)}")
    
    # Check Category Summary
    ws_cat = wb["Category Summary"]
    print(f"  Category Summary max_row: {ws_cat.max_row} (expected 21)")
    bad_cats_cat = []
    for r in range(3, ws_cat.max_row):
        cat = ws_cat.cell(row=r, column=1).value
        if cat in ["Construction & Hardware", "Automobiles & Accessories"]:
            bad_cats_cat.append((r, cat))
    print(f"  Category Summary bad categories found: {len(bad_cats_cat)}")
    
    # Check Daily Totals
    ws_dt = wb["Daily Totals"]
    print(f"  Daily Totals columns count: {ws_dt.max_column} (expected 7)")
    headers = [ws_dt.cell(row=2, column=c).value for c in range(1, ws_dt.max_column + 1)]
    print(f"  Daily Totals headers: {headers}")
    
    # Check Financial Summary
    ws_fs = wb["Financial Summary"]
    b6 = ws_fs["B6"].value
    b7 = ws_fs["B7"].value
    b22 = ws_fs["B22"].value
    b23 = ws_fs["B23"].value
    b24 = ws_fs["B24"].value
    b25 = ws_fs["B25"].value
    print(f"  Financial Summary B6 (Revenue): {b6} (expected ='Revenue Analysis'!F164)")
    print(f"  Financial Summary B7 (COGS): {b7} (expected ='Revenue Analysis'!G164)")
    print(f"  Financial Summary B22 (Closing Stock): {b22} (expected =SUMPRODUCT('Sales Data'!G3:G163,'Sales Data'!K3:K163))")
    print(f"  Financial Summary B23 (SKU count): {b23} (expected =COUNTA('Sales Data'!A3:A163))")
    print(f"  Financial Summary B24 (Low Stock): {b24} (expected =COUNTIF('Sales Data'!N3:N163,\"Low Stock\"))")
    print(f"  Financial Summary B25 (Avg Daily): {b25} (expected ='Revenue Analysis'!E164/30 or similar)")
    
    # Check Revenue Analysis
    ws_ra = wb["Revenue Analysis"]
    print(f"  Revenue Analysis max_row: {ws_ra.max_row} (expected 164)")
    totals_val = ws_ra.cell(row=ws_ra.max_row, column=1).value
    print(f"  Revenue Analysis row {ws_ra.max_row} Col 1: {totals_val} (expected TOTALS)")
    
    wb.close()

verify_file("AlNoor_Financial_Summary_April_2026.xlsx")
