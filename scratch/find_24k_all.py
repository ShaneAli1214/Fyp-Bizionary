import os
import openpyxl

for filename in os.listdir('output'):
    if not filename.endswith('.xlsx') or filename.startswith('~$'):
        continue
    filepath = os.path.join('output', filename)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        for name in wb.sheetnames:
            s = wb[name]
            # check the first few rows and columns
            for r in range(1, min(s.max_row+1, 220)):
                for c in range(1, min(s.max_column+1, 50)):
                    val = s.cell(row=r, column=c).value
                    if val is not None:
                        val_str = str(val).lower()
                        if '24000' in val_str or '24k' in val_str or '24,' in val_str:
                            # print if it looks like a total or category total
                            print(f"{filename} | Sheet: {name} | Cell: {openpyxl.utils.get_column_letter(c)}{r} | Value: {val}")
        wb.close()
    except Exception as e:
        print(f"Error {filename}: {e}")
