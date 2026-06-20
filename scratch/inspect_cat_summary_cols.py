import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Category Summary"]
print("Header Row 2:")
print([sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)])
print("Row 28 (GRAND TOTAL):")
print([sheet.cell(row=28, column=c).value for c in range(1, sheet.max_column + 1)])
wb.close()
