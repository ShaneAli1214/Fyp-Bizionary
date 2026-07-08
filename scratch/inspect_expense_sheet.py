import openpyxl

wb = openpyxl.load_workbook("Expense_Sheet.xlsx", data_only=True)
sheet = wb['Monthly Expenses']
print(f"Total rows: {sheet.max_row}, Total columns: {sheet.max_column}")
for r in range(1, min(60, sheet.max_row + 1)):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
    # Skip printing if the whole row is empty
    if any(val is not None for val in row_vals):
        print(f"Row {r:02d}: {row_vals}")
wb.close()
