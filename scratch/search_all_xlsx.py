import os
import openpyxl

files_to_check = []
# Check root folder files
for file in os.listdir("."):
    if file.endswith(".xlsx") and not file.startswith("~$"):
        files_to_check.append(file)

# Check output/ folder files
if os.path.exists("output"):
    for file in os.listdir("output"):
        if file.endswith(".xlsx") and not file.startswith("~$"):
            files_to_check.append(os.path.join("output", file))

print("Files to check:", files_to_check)

for path in files_to_check:
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        has_hardware = False
        has_automobile = False
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            # iterate rows quickly in read-only mode
            for row in ws.iter_rows(max_row=300, max_col=15, values_only=True):
                for val in row:
                    if val:
                        val_str = str(val).lower()
                        if "hardware" in val_str:
                            has_hardware = True
                        if "automobile" in val_str:
                            has_automobile = True
        wb.close()
        if has_hardware or has_automobile:
            print(f"File {path}: has_hardware={has_hardware}, has_automobile={has_automobile}")
    except Exception as e:
        print(f"Error reading {path}: {e}")
