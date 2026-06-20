import openpyxl

wb = openpyxl.load_workbook('output/30day_sales_AlNoor_cleaned.xlsx')
s = wb['Category Summary']
for r in range(1, s.max_row+1):
    vals = [s.cell(r, c).value for c in range(1, 9)]
    if any(vals):
        print(f"Row {r}: {vals}")
wb.close()
