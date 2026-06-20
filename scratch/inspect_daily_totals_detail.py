import openpyxl

wb = openpyxl.load_workbook("output/AlNoor_April_2026.xlsx")
sheet = wb["Daily Totals"]
print("Row 3 detail:")
for c in range(1, sheet.max_column + 1):
    print(f"Col {c}: value={sheet.cell(row=3, column=c).value}, type={type(sheet.cell(row=3, column=c).value)}")
wb.close()
