import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Sales Data"]
print("Row | SKU | Product Name | Category | Subcategory")
for r in range(3, sheet.max_row + 1):
    sku = sheet.cell(row=r, column=2).value
    name = sheet.cell(row=r, column=1).value
    cat = sheet.cell(row=r, column=3).value
    subcat = sheet.cell(row=r, column=4).value
    if cat in ["Construction & Hardware", "Automobiles & Accessories"]:
        print(f"{r} | {sku} | {name} | {cat} | {subcat}")
wb.close()
