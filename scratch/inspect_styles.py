import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Financial Summary"]

print("=== Financial Summary Styles ===")
for r in range(1, 20):
    cell = sheet.cell(row=r, column=1)
    fill = cell.fill
    font = cell.font
    fill_color = fill.start_color.rgb if fill and fill.fill_type else "None"
    font_bold = font.bold if font else False
    font_color = font.color.rgb if font and font.color else "None"
    print(f"Row {r}, Col 1: Val='{cell.value}', Bold={font_bold}, FontColor={font_color}, FillColor={fill_color}")

# Let's inspect column widths too
print("\n=== Column Widths ===")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    print(f"Col {col}: width={sheet.column_dimensions[col].width}")
