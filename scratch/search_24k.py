import openpyxl

wb = openpyxl.load_workbook('AlNoor_Financial_Summary.xlsx')
for name in wb.sheetnames:
    s = wb[name]
    for r in range(1, s.max_row+1):
        for c in range(1, s.max_column+1):
            val = s.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).lower()
                if '24000' in val_str or '24k' in val_str or '24,000' in val_str:
                    print(f"Sheet: {name} | Cell: {openpyxl.utils.get_column_letter(c)}{r} | Value: {val}")
wb.close()
