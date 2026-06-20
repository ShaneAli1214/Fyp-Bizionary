import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx", data_only=False)

print("=== Revenue Analysis Formulas ===")
sheet_ra = wb["Revenue Analysis"]
for r in [1, 2, 3, 4, 215, 216, 217]:
    row_vals = [sheet_ra.cell(row=r, column=c).value for c in range(1, 9)]
    print(f"Row {r}: {row_vals}")

print("\n=== Cash Flow Statement Formulas ===")
sheet_cf = wb["Cash Flow Statement"]
for r in [1, 2, 3, 4, 73, 74, 75, 76, 77, 78]:
    row_vals = [sheet_cf.cell(row=r, column=c).value for c in range(1, 7)]
    print(f"Row {r}: {row_vals}")
