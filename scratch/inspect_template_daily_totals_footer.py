import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Daily Totals"]
print("Max row:", sheet.max_row)
for r in range(max(1, sheet.max_row - 5), sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")
wb.close()
