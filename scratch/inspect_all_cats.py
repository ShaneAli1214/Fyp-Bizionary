import openpyxl

wb = openpyxl.load_workbook("output/AlNoor_April_2026.xlsx")
sheet = wb["Category Summary"]
print("Category | Sub-category | # Prods | Tot Sold | Avg/Prod | Min | Max | Avg Daily")
for r in range(3, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 9)]
    print(vals)
wb.close()
