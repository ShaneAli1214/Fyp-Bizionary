import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Revenue Analysis"]
print("Header Row 2:")
print([sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)])
print("Row 3 sample:")
for c in range(1, sheet.max_column + 1):
    print(f"Col {c}: value={sheet.cell(row=3, column=c).value}, type={type(sheet.cell(row=3, column=c).value)}")
wb.close()
