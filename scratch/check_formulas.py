import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx", data_only=False)
sheet = wb["Financial Summary"]

print("=== Financial Summary Formulas (data_only=False) ===")
for r in [5, 6, 7, 8, 9, 10, 11, 12, 15, 16, 17, 18, 21, 22, 23, 24]:
    cell_val = sheet.cell(row=r, column=2).value
    print(f"Row {r}: '{cell_val}'")

sheet_kpi = wb["KPI Dashboard Data"]
print("\n=== KPI Dashboard Data Formulas ===")
for r in range(2, 18):
    cell_val = sheet_kpi.cell(row=r, column=2).value
    print(f"Row {r}: '{cell_val}'")
