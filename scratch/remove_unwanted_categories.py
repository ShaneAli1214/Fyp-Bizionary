import os
import shutil
import openpyxl

files_to_process = [
    "AlNoor_Financial_Summary.xlsx",
    "output/AlNoor_January_2026.xlsx",
    "output/AlNoor_February_2026.xlsx",
    "output/AlNoor_March_2026.xlsx",
    "output/AlNoor_April_2026.xlsx",
    "output/AlNoor_June_2026.xlsx",
    "output/30day_sales_AlNoor_cleaned.xlsx"
]

unwanted_categories = ["Construction & Hardware", "Automobiles & Accessories"]

def process_workbook(path):
    print(f"\nProcessing workbook: {path}")
    
    # 1. Create backup
    bak_path = path + ".bak"
    if not os.path.exists(bak_path):
        shutil.copy2(path, bak_path)
        print(f"  Backup created at {bak_path}")
    
    # 2. Load workbook
    wb = openpyxl.load_workbook(path)
    
    # --- SALES DATA SHEET ---
    if "Sales Data" in wb.sheetnames:
        print("  Processing Sales Data...")
        ws = wb["Sales Data"]
        # Delete unwanted categories from rows
        # Loop from bottom to top (row 3 to max_row)
        rows_deleted = 0
        for r in range(ws.max_row, 2, -1):
            cat_val = ws.cell(row=r, column=3).value
            if cat_val in unwanted_categories:
                ws.delete_rows(r, 1)
                rows_deleted += 1
        print(f"  Deleted {rows_deleted} product rows from Sales Data. Remaining rows: {ws.max_row}")

    # --- REVENUE ANALYSIS SHEET ---
    if "Revenue Analysis" in wb.sheetnames:
        print("  Processing Revenue Analysis...")
        ws = wb["Revenue Analysis"]
        rows_deleted = 0
        for r in range(ws.max_row, 2, -1):
            cat_val = ws.cell(row=r, column=3).value
            if cat_val in unwanted_categories:
                ws.delete_rows(r, 1)
                rows_deleted += 1
        print(f"  Deleted {rows_deleted} product rows from Revenue Analysis. Remaining rows: {ws.max_row}")
        
        # Recalculate totals row at row 164 (since 217 - 53 = 164)
        totals_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "TOTALS":
                totals_row = r
                break
        if totals_row:
            print(f"  Found TOTALS row at row {totals_row}")
            ws.cell(row=totals_row, column=5).value = f"=SUM(E3:E{totals_row-1})"
            ws.cell(row=totals_row, column=6).value = f"=SUM(F3:F{totals_row-1})"
            ws.cell(row=totals_row, column=7).value = f"=SUM(G3:G{totals_row-1})"
            ws.cell(row=totals_row, column=8).value = f"=SUM(H3:H{totals_row-1})"
            
    # --- CATEGORY SUMMARY SHEET ---
    if "Category Summary" in wb.sheetnames:
        print("  Processing Category Summary...")
        ws = wb["Category Summary"]
        rows_deleted = 0
        for r in range(ws.max_row, 2, -1):
            cat_val = ws.cell(row=r, column=1).value
            if cat_val in unwanted_categories:
                ws.delete_rows(r, 1)
                rows_deleted += 1
        print(f"  Deleted {rows_deleted} category/subcategory rows from Category Summary. Remaining rows: {ws.max_row}")
        
        grand_total_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "GRAND TOTAL":
                grand_total_row = r
                break
        if grand_total_row:
            print(f"  Found GRAND TOTAL row at row {grand_total_row}")
            # If it's a template or output summary, set the formula or value
            if "Financial_Summary.xlsx" in path or "Financial_Summary_" in path:
                ws.cell(row=grand_total_row, column=4).value = f"=SUM(D3:D{grand_total_row-1})"
                ws.cell(row=grand_total_row, column=5).value = f"=AVERAGE(E3:E{grand_total_row-1})"
            else:
                # In output files, they are hardcoded integers
                sum_val = sum(ws.cell(row=r, column=4).value or 0 for r in range(3, grand_total_row))
                ws.cell(row=grand_total_row, column=4).value = sum_val
                print(f"  Recalculated GRAND TOTAL Units Sold to {sum_val}")
                
    # --- DAILY TOTALS SHEET ---
    if "Daily Totals" in wb.sheetnames:
        print("  Processing Daily Totals...")
        ws = wb["Daily Totals"]
        # Find columns to delete in row 2
        cols_to_delete = []
        for c in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=c).value
            if header in unwanted_categories:
                cols_to_delete.append(c)
        
        # Delete columns in reverse order to keep indices stable
        cols_to_delete.sort(reverse=True)
        for c in cols_to_delete:
            ws.delete_cols(c, 1)
            print(f"  Deleted column index {c} from Daily Totals")
            
        # Re-identify Daily totals columns and Day Total
        day_total_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=2, column=c).value == "Day Total":
                day_total_col = c
                break
                
        if day_total_col:
            # Update Day Total values/formulas
            is_template_or_summary = "Financial_Summary.xlsx" in path or "Financial_Summary_" in path
            for r in range(3, ws.max_row + 1):
                is_total_row = (ws.cell(row=r, column=1).value == "TOTAL")
                
                if is_total_row:
                    if is_template_or_summary:
                        # Write formulas for the TOTAL row
                        ws.cell(row=r, column=2).value = f"=SUM(B3:B{r-1})"
                        ws.cell(row=r, column=3).value = f"=SUM(C3:C{r-1})"
                        ws.cell(row=r, column=4).value = f"=SUM(D3:D{r-1})"
                        ws.cell(row=r, column=5).value = f"=SUM(E3:E{r-1})"
                        ws.cell(row=r, column=6).value = f"=SUM(F3:F{r-1})"
                        ws.cell(row=r, column=7).value = f"=SUM(G3:G{r-1})"
                    else:
                        # Write hardcoded sums for the TOTAL row (input files)
                        for c in range(2, 7):
                            col_sum = sum(ws.cell(row=temp_r, column=c).value or 0 for temp_r in range(3, r))
                            ws.cell(row=r, column=c).value = col_sum
                        day_sum = sum(ws.cell(row=r, column=cat_c).value or 0 for cat_c in range(2, 7))
                        ws.cell(row=r, column=7).value = day_sum
                else:
                    if is_template_or_summary:
                        ws.cell(row=r, column=day_total_col).value = f"=SUM(B{r}:F{r})"
                    else:
                        day_sum = sum(ws.cell(row=r, column=cat_c).value or 0 for cat_c in range(2, day_total_col))
                        ws.cell(row=r, column=day_total_col).value = day_sum
            print(f"  Recalculated Day Total values in column {day_total_col}")
            
    # --- FINANCIAL SUMMARY SHEET ---
    if "Financial Summary" in wb.sheetnames:
        print("  Processing Financial Summary...")
        ws = wb["Financial Summary"]
        # Row 2 Products count
        val_2 = ws.cell(row=2, column=1).value
        if val_2 and "214 Active SKUs" in val_2:
            ws.cell(row=2, column=1).value = val_2.replace("214 Active SKUs", "161 Active SKUs")
        
        # Row 6: Total Revenue formula and description
        ws.cell(row=6, column=2).value = "='Revenue Analysis'!F164"
        note_6 = ws.cell(row=6, column=3).value
        if note_6 and "214 products" in note_6:
            ws.cell(row=6, column=3).value = note_6.replace("214 products", "161 products")
            
        # Row 7: Total COGS formula and description
        ws.cell(row=7, column=2).value = "='Revenue Analysis'!G164"
        note_7 = ws.cell(row=7, column=3).value
        if note_7 and "214 products" in note_7:
            ws.cell(row=7, column=3).value = note_7.replace("214 products", "161 products")
            
        # Row 22: Closing Stock Value formula
        ws.cell(row=22, column=2).value = "=SUMPRODUCT('Sales Data'!G3:G163,'Sales Data'!K3:K163)"
        
        # Row 23: Total Active SKUs formula and description
        ws.cell(row=23, column=2).value = "=COUNTA('Sales Data'!A3:A163)"
        note_23 = ws.cell(row=23, column=3).value
        if note_23 and "214 SKUs across 6 major categories" in note_23:
            ws.cell(row=23, column=3).value = note_23.replace("214 SKUs across 6 major categories", "161 SKUs across 5 major categories")
            
        # Row 24: Low Stock Products formula
        ws.cell(row=24, column=2).value = "=COUNTIF('Sales Data'!N3:N163,\"Low Stock\")"
        
        # Row 25: Average Daily Sales Volume formula
        ws.cell(row=25, column=2).value = "='Revenue Analysis'!E164/30"
        
    # --- KPI DASHBOARD DATA SHEET ---
    if "KPI Dashboard Data" in wb.sheetnames:
        print("  Processing KPI Dashboard Data...")
        ws = wb["KPI Dashboard Data"]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    if "214" in val:
                        ws.cell(row=r, column=c).value = val.replace("214", "161")
                        
    wb.save(path)
    wb.close()
    print(f"  Workbook successfully saved: {path}")

for p in files_to_process:
    if os.path.exists(p):
        process_workbook(p)
    else:
        print(f"Workbook not found: {p}")

print("\nCleanup completed.")
