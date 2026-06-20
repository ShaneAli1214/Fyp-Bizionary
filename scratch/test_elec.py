import openpyxl

wb = openpyxl.load_workbook('output/AlNoor_February_2026.xlsx')
s = wb['Sales Data']
print("Cols:", [s.cell(2, c).value for c in range(1, 15)])

total_units = 0
total_rev = 0

for r in range(3, s.max_row+1):
    cat = s.cell(row=r, column=3).value
    subcat = s.cell(row=r, column=4).value
    if cat == 'Electronics & Appliances':
        # Daily sales start at col 13
        row_units = sum(int(s.cell(row=r, column=c).value or 0) for c in range(13, s.max_column+1))
        total_units += row_units
        price_val = s.cell(row=r, column=11).value
        # remove commas
        price = float(str(price_val).replace(',', ''))
        total_rev += row_units * price
        
        # Check stock values
        qty_in_hand = s.cell(row=r, column=7).value
        stock_status = s.cell(row=r, column=12).value
        print(f"Product: {s.cell(row=r, column=2).value} ({subcat}) | Unit Price: {price} | Units Sold: {row_units} | Qty In Hand: {qty_in_hand} | Status: {stock_status}")

print("Total units sold in Electronics:", total_units)
print("Total revenue in Electronics:", total_rev)
wb.close()
