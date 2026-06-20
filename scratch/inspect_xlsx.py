import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary_April_2026.xlsx")
print("Sheets:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    found = False
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and any(word in str(val).lower() for word in ["hardware", "automobile"]):
                print(f"Found in sheet '{sheetname}', row {r}, col {c} ({openpyxl.utils.get_column_letter(c)}{r}): {val}")
                found = True
wb.close()
