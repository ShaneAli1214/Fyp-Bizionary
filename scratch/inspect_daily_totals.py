import openpyxl

wb = openpyxl.load_workbook("output/AlNoor_April_2026.xlsx")
sheet = wb["Daily Totals"]
print("Header row 1:")
print([sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)])
print("Header row 2:")
print([sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)])
wb.close()
