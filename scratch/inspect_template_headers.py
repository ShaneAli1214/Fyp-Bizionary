import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"Sheet: {name}, Cell A1: {ws['A1'].value}")
wb.close()
