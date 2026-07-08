import os
import sys
import django
from datetime import date
from decimal import Decimal
import calendar
import openpyxl

sys.path.append(os.getcwd())
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp_system.settings')
django.setup()

from accounts.models import UtilityBill, SalaryPayment, RecurringCost, Expense, CashTransaction, JournalEntry
from user_management.models import ERPUser

def main():
    print("Clearing existing financial data...")
    # Clear child objects first by deleting parents
    UtilityBill.objects.all().delete()
    SalaryPayment.objects.all().delete()
    RecurringCost.objects.all().delete()
    # Delete any other expenses (except supplies created from slips)
    Expense.objects.exclude(category='SUPPLIES').delete()
    print("Data cleared!")

    wb = openpyxl.load_workbook("Expense_Sheet.xlsx", data_only=True)
    sheet = wb['Monthly Expenses']

    users = {
        'Finance Accountant': ERPUser.objects.get(username='accountant'),
        'Sales Manager': ERPUser.objects.get(username='sales'),
        'Inventory Manager': ERPUser.objects.get(username='inventory'),
    }

    # Helper function to get last day of month
    def get_last_day(year, month):
        return calendar.monthrange(year, month)[1]

    # Import Utility Bills
    # Row 5 (Electricity), Row 6 (Gas), Row 7 (Water), Row 8 (Internet)
    utility_rows = {
        5: 'ELECTRICITY',
        6: 'GAS',
        7: 'WATER',
        8: 'INTERNET'
    }

    print("\nImporting Utility Bills...")
    for row_idx, utype in utility_rows.items():
        name = sheet.cell(row=row_idx, column=1).value
        print(f"Processing row {row_idx}: {name} ({utype})")
        for m in range(1, 13):
            val = sheet.cell(row=row_idx, column=m + 1).value
            if val is not None:
                amount = Decimal(str(val))
                last_day = get_last_day(2026, m)
                bill_date = date(2026, m, 1)
                
                UtilityBill.objects.create(
                    utility_type=utype,
                    bill_number=f"BILL-2026-{m:02d}-{utype[:3]}",
                    billing_period_start=date(2026, m, 1),
                    billing_period_end=date(2026, m, last_day),
                    due_date=bill_date,
                    amount=amount,
                    tax_amount=Decimal('0.00'),
                    payment_date=bill_date,
                    payment_method='CASH',
                    status='PAID',
                    notes=f"Imported monthly {name.lower()} bill for {bill_date.strftime('%B %Y')}"
                )

    # Import Salaries
    # Row 12 (Finance Accountant), Row 13 (Sales Manager), Row 14 (Inventory Manager)
    salary_rows = {
        12: 'Finance Accountant',
        13: 'Sales Manager',
        14: 'Inventory Manager'
    }

    print("\nImporting Salaries...")
    for row_idx, employee_title in salary_rows.items():
        name = sheet.cell(row=row_idx, column=1).value
        print(f"Processing row {row_idx}: {name}")
        user = users[employee_title]
        for m in range(1, 13):
            val = sheet.cell(row=row_idx, column=m + 1).value
            if val is not None:
                amount = Decimal(str(val))
                last_day = get_last_day(2026, m)
                pay_date = date(2026, m, 1)
                
                SalaryPayment.objects.create(
                    employee=user,
                    amount=amount,
                    pay_period_start=date(2026, m, 1),
                    pay_period_end=date(2026, m, last_day),
                    payment_date=pay_date,
                    payment_method='BANK_TRANSFER',
                    status='PAID',
                    notes=f"Imported salary for {name} - {pay_date.strftime('%B %Y')}"
                )

    # Import Recurring Operational Costs
    # Row 18 (Warehouse Management Fee), Row 19 (Handling & Storage Charges)
    recurring_rows = {
        18: ('RENT', 'Warehouse Management Fee'),
        19: ('OTHER', 'Handling & Storage Charges')
    }

    print("\nImporting Recurring Costs...")
    for row_idx, (ctype, cost_name) in recurring_rows.items():
        name = sheet.cell(row=row_idx, column=1).value
        print(f"Processing row {row_idx}: {name} ({ctype})")
        for m in range(1, 13):
            val = sheet.cell(row=row_idx, column=m + 1).value
            if val is not None:
                amount = Decimal(str(val))
                due_date = date(2026, m, 1)
                
                RecurringCost.objects.create(
                    cost_type=ctype,
                    name=cost_name,
                    amount=amount,
                    due_date=due_date,
                    payment_date=due_date,
                    payment_method='BANK_TRANSFER',
                    status='PAID',
                    notes=f"Imported {cost_name} for {due_date.strftime('%B %Y')}"
                )

    print("\nData import completed successfully!")
    wb.close()

if __name__ == '__main__':
    main()
