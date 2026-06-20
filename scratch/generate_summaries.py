import os
import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = Path("c:/Users/Dell/Desktop/Fyp")
OUTPUT_DIR = PROJECT_ROOT / "output"
TEMPLATE_PATH = PROJECT_ROOT / "AlNoor_Financial_Summary.xlsx"
DATASET_PATH = PROJECT_ROOT / "AlNoor_ERP_Dataset.xlsx"

# 1. Build product lookup from AlNoor_ERP_Dataset.xlsx
print("Loading product master catalog...")
dataset_wb = openpyxl.load_workbook(DATASET_PATH, read_only=True)
products_sheet = dataset_wb["Products"]

# We need to find columns for Product ID and Cost Price (PKR)
# Row 1 is the title row, Row 2 is the header row
headers = [cell.value for cell in products_sheet[2]]
product_id_idx = headers.index("Product ID") + 1
cost_price_idx = headers.index("Cost Price (PKR)") + 1
sale_price_idx = headers.index("Sale Price (PKR)") + 1

product_cost_map = {}
# Product records start at row 3
for r in range(3, 164):
    pid = products_sheet.cell(row=r, column=product_id_idx).value
    cost = products_sheet.cell(row=r, column=cost_price_idx).value
    if pid:
        product_cost_map[str(pid).strip()] = cost

dataset_wb.close()
print(f"Loaded {len(product_cost_map)} products from master catalog.")

def get_column_letter(col_idx):
    """Convert a 1-based column index to Excel column letter (e.g. 1 -> A, 27 -> AA)"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

def copy_cell(src_cell, dest_cell):
    """Copy value and formatting from src_cell to dest_cell"""
    dest_cell.value = src_cell.value
    if src_cell.has_style:
        dest_cell.font = copy.copy(src_cell.font)
        dest_cell.fill = copy.copy(src_cell.fill)
        dest_cell.border = copy.copy(src_cell.border)
        dest_cell.alignment = copy.copy(src_cell.alignment)
        dest_cell.number_format = src_cell.number_format

def clear_sheet(sheet):
    """Clear all cells and formatting from a sheet"""
    sheet.delete_rows(1, sheet.max_row + 10)
    sheet.delete_cols(1, sheet.max_column + 10)

def generate_monthly_summary(month_name, year_val, input_filename, output_filename):
    print(f"\nGenerating financial summary for {month_name} {year_val}...")
    # Load template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # Load input data
    input_path = OUTPUT_DIR / input_filename
    in_wb = openpyxl.load_workbook(input_path)
    
    # -------------------------------------------------------------
    # 1. Update Sales Data Sheet
    # -------------------------------------------------------------
    print("  Processing Sales Data...")
    in_sales_sheet = in_wb["Sales Data"]
    temp_sales_sheet = wb["Sales Data"]
    clear_sheet(temp_sales_sheet)
    
    # Copy column dimensions (widths)
    for col_col, col_dim in in_sales_sheet.column_dimensions.items():
        temp_sales_sheet.column_dimensions[col_col] = copy.copy(col_dim)
        
    num_rows = in_sales_sheet.max_row
    num_cols = in_sales_sheet.max_column
    
    # We will copy input sales data and insert Cost Price (K) and Gross Marjin (M)
    for r in range(1, num_rows + 1):
        # A to J (1 to 10)
        for c in range(1, 11):
            src_c = in_sales_sheet.cell(row=r, column=c)
            dest_c = temp_sales_sheet.cell(row=r, column=c)
            copy_cell(src_c, dest_c)
            
        # K (Cost Price)
        dest_k = temp_sales_sheet.cell(row=r, column=11)
        if r == 1:
            dest_k.value = in_sales_sheet.cell(row=r, column=1).value # Title row
        elif r == 2:
            dest_k.value = "Cost Price"
            # Style it like the header
            src_header = in_sales_sheet.cell(row=2, column=10) # Stock Value header as style source
            dest_k.font = copy.copy(src_header.font)
            dest_k.fill = copy.copy(src_header.fill)
            dest_k.border = copy.copy(src_header.border)
            dest_k.alignment = copy.copy(src_header.alignment)
        else:
            pid = temp_sales_sheet.cell(row=r, column=1).value
            dest_k.value = product_cost_map.get(str(pid).strip(), 0.0)
            # Style it like other product fields (numeric)
            src_style = in_sales_sheet.cell(row=r, column=10) # Stock Value cell as style source
            dest_k.font = copy.copy(src_style.font)
            dest_k.fill = copy.copy(src_style.fill)
            dest_k.border = copy.copy(src_style.border)
            dest_k.alignment = copy.copy(src_style.alignment)
            dest_k.number_format = '#,##0.00'
            
        # L (Sale Price - originally Column K (11))
        for c in range(11, 12):
            src_c = in_sales_sheet.cell(row=r, column=c)
            dest_c = temp_sales_sheet.cell(row=r, column=12)
            copy_cell(src_c, dest_c)
            
        # M (Gross Marjin)
        dest_m = temp_sales_sheet.cell(row=r, column=13)
        if r == 1:
            dest_m.value = in_sales_sheet.cell(row=r, column=1).value
        elif r == 2:
            dest_m.value = "Gross Marjin"
            src_header = in_sales_sheet.cell(row=2, column=11) # Sale Price header as style source
            dest_m.font = copy.copy(src_header.font)
            dest_m.fill = copy.copy(src_header.fill)
            dest_m.border = copy.copy(src_header.border)
            dest_m.alignment = copy.copy(src_header.alignment)
        else:
            dest_m.value = f"=IF(L{r}=0,0,(L{r}-K{r})/L{r})"
            src_style = in_sales_sheet.cell(row=r, column=11)
            dest_m.font = copy.copy(src_style.font)
            dest_m.fill = copy.copy(src_style.fill)
            dest_m.border = copy.copy(src_style.border)
            dest_m.alignment = copy.copy(src_style.alignment)
            dest_m.number_format = '0.0%'
            
        # N (Stock Status - originally Column L (12))
        for c in range(12, 13):
            src_c = in_sales_sheet.cell(row=r, column=c)
            dest_c = temp_sales_sheet.cell(row=r, column=14)
            copy_cell(src_c, dest_c)
            
        # O onwards (Daily Sales - originally Column M (13) onwards)
        for c in range(13, num_cols + 1):
            src_c = in_sales_sheet.cell(row=r, column=c)
            dest_c = temp_sales_sheet.cell(row=r, column=c + 2)
            copy_cell(src_c, dest_c)
            
    # Set the Title of Sales Data
    temp_sales_sheet.cell(row=1, column=1).value = f"Al-Noor Trading — 30-Day Sales Data  ({month_name} 1 – {month_name} {num_cols - 12})"
    
    # -------------------------------------------------------------
    # 2. Update Category Summary Sheet
    # -------------------------------------------------------------
    print("  Processing Category Summary...")
    in_cat_sheet = in_wb["Category Summary"]
    temp_cat_sheet = wb["Category Summary"]
    clear_sheet(temp_cat_sheet)
    for r in range(1, in_cat_sheet.max_row + 1):
        for c in range(1, in_cat_sheet.max_column + 1):
            copy_cell(in_cat_sheet.cell(row=r, column=c), temp_cat_sheet.cell(row=r, column=c))
            
    # -------------------------------------------------------------
    # 3. Update Daily Totals Sheet
    # -------------------------------------------------------------
    print("  Processing Daily Totals...")
    in_dt_sheet = in_wb["Daily Totals"]
    temp_dt_sheet = wb["Daily Totals"]
    clear_sheet(temp_dt_sheet)
    for r in range(1, in_dt_sheet.max_row + 1):
        for c in range(1, in_dt_sheet.max_column + 1):
            copy_cell(in_dt_sheet.cell(row=r, column=c), temp_dt_sheet.cell(row=r, column=c))
            
    # -------------------------------------------------------------
    # 4. Update Expense Tracker Sheet
    # -------------------------------------------------------------
    print("  Processing Expense Tracker...")
    temp_exp_sheet = wb["Expense Tracker"]
    # Change sheet title
    temp_exp_sheet.cell(row=1, column=1).value = f"Al-Noor Trading — Expense Tracker ({month_name} {year_val})"
    
    # We shift dates in Column A (Date). Row 3 to 15 (13 records)
    # The day numbers: 1, 2, 4, 6, 8, 10, 12, 15, 18, 20, 23, 27 (active), 14 (voided)
    day_numbers = [1, 2, 4, 6, 8, 10, 12, 15, 18, 20, 23, 27, 14]
    month_num = datetime_month_number(month_name)
    
    for idx, r in enumerate(range(3, 16)):
        day_num = day_numbers[idx]
        date_str = f"{year_val}-{month_num:02d}-{day_num:02d}"
        temp_exp_sheet.cell(row=r, column=1).value = date_str
        
    # Update total text and voided notes if any
    temp_exp_sheet.cell(row=16, column=1).value = f"TOTAL ACTIVE EXPENSES (PKR)  — 12 records, no tax"
    temp_exp_sheet.cell(row=17, column=1).value = f"1 VOIDED record ({year_val}-{month_num:02d}-14, Grey, PKR 1,500.00) excluded from active totals"
    
    # -------------------------------------------------------------
    # 5. Update Revenue Analysis Sheet
    # -------------------------------------------------------------
    print("  Processing Revenue Analysis...")
    temp_ra_sheet = wb["Revenue Analysis"]
    temp_ra_sheet.cell(row=1, column=1).value = f"Al-Noor Trading — Revenue Analysis ({month_name} {year_val})"
    
    # Daily columns in Sales Data sheet are O to AS (index 15 to 45 max)
    num_days = num_cols - 12
    end_col_letter = get_column_letter(14 + num_days) # Column O is 15th (1-indexed)
    
    for r in range(3, 164):
        # Update formula for Qty Sold (Col E)
        temp_ra_sheet.cell(row=r, column=5).value = f"=SUMPRODUCT('Sales Data'!O{r}:{end_col_letter}{r})"
        # Update formula for Total Revenue (Col F)
        temp_ra_sheet.cell(row=r, column=6).value = f"=SUMPRODUCT('Sales Data'!O{r}:{end_col_letter}{r})*'Sales Data'!L{r}"
        # Update formula for Total COGS (Col G)
        temp_ra_sheet.cell(row=r, column=7).value = f"=SUMPRODUCT('Sales Data'!O{r}:{end_col_letter}{r})*'Sales Data'!K{r}"
        
    temp_ra_sheet.cell(row=164, column=5, value="=SUM(E3:E163)")
    
    # -------------------------------------------------------------
    # 6. Update Cash Flow Statement Sheet
    # -------------------------------------------------------------
    print("  Processing Cash Flow Statement...")
    temp_cf_sheet = wb["Cash Flow Statement"]
    clear_sheet(temp_cf_sheet)
    
    # Copy column dimensions
    temp_cf_sheet.column_dimensions['A'].width = 15.0
    temp_cf_sheet.column_dimensions['B'].width = 10.0
    temp_cf_sheet.column_dimensions['C'].width = 12.0
    temp_cf_sheet.column_dimensions['D'].width = 34.0
    temp_cf_sheet.column_dimensions['E'].width = 15.0
    temp_cf_sheet.column_dimensions['F'].width = 16.0
    
    # Style template components (borders, fonts, fills)
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    fill_title = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")
    
    font_header = Font(name="Calibri", size=11, bold=True, color="000000")
    fill_header = PatternFill(fill_type="solid", start_color="9DC3E6", end_color="9DC3E6")
    
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, bold=False, color="000000")
    fill_summary = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    border_summary = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # Write Title
    temp_cf_sheet.merge_cells("A1:F1")
    title_cell = temp_cf_sheet.cell(row=1, column=1)
    title_cell.value = f"Al-Noor Trading — Cash Flow Statement ({month_name} {year_val})"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    temp_cf_sheet.row_dimensions[1].height = 35.0
    
    # Write Headers
    headers_cf = ['Date', 'Day', 'Type', 'Description', 'Reference', 'Amount (PKR)']
    for c, h in enumerate(headers_cf, 1):
        cell = temp_cf_sheet.cell(row=2, column=c)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if c != 4 else "left", vertical="center")
        cell.border = border_thin
    temp_cf_sheet.row_dimensions[2].height = 24.0
    
    import datetime
    
    cf_row = 3
    # Expenses active mapping
    expense_data = [
        {"day": 1, "desc": "Office stationery & supplies", "ref": "EXP-05-01", "amt": 2500},
        {"day": 2, "desc": "Packaging materials", "ref": "EXP-05-02", "amt": 3200},
        {"day": 4, "desc": "Cleaning & hygiene supplies", "ref": "EXP-05-04", "amt": 1800},
        {"day": 6, "desc": "Display & shelving items", "ref": "EXP-05-06", "amt": 4500},
        {"day": 8, "desc": "Printer paper & ink cartridges", "ref": "EXP-05-08", "amt": 2750},
        {"day": 10, "desc": "Storage boxes & labels", "ref": "EXP-05-10", "amt": 1950},
        {"day": 12, "desc": "Wrapping & packing tape", "ref": "EXP-05-12", "amt": 1200},
        {"day": 15, "desc": "Security & safety supplies", "ref": "EXP-05-15", "amt": 3400},
        {"day": 18, "desc": "Promotional signage material", "ref": "EXP-05-18", "amt": 2800},
        {"day": 20, "desc": "Counter & billing supplies", "ref": "EXP-05-20", "amt": 3589.99},
        {"day": 23, "desc": "Inventory tags & barcodes", "ref": "EXP-05-23", "amt": 2100},
        {"day": 27, "desc": "General office consumables", "ref": "EXP-05-27", "amt": 1800},
    ]
    
    for day in range(1, num_days + 1):
        date_obj = datetime.date(year_val, month_num, day)
        day_name = date_obj.strftime("%a")
        date_str = date_obj.isoformat()
        
        sales_col_letter = get_column_letter(14 + day) # Col O is 15th
        
        # 1. INFLOW row
        temp_cf_sheet.cell(row=cf_row, column=1, value=date_str)
        temp_cf_sheet.cell(row=cf_row, column=2, value=day_name)
        temp_cf_sheet.cell(row=cf_row, column=3, value="INFLOW")
        temp_cf_sheet.cell(row=cf_row, column=4, value="Daily Sales Revenue")
        temp_cf_sheet.cell(row=cf_row, column=5, value=f"REV-{month_num:02d}-{day:02d}")
        cell_amt = temp_cf_sheet.cell(row=cf_row, column=6, value=f"=SUMPRODUCT('Sales Data'!{sales_col_letter}3:{sales_col_letter}163,'Sales Data'!$L$3:$L$163)")
        cell_amt.number_format = '#,##0.00'
        
        for c in range(1, 7):
            temp_cf_sheet.cell(row=cf_row, column=c).border = border_thin
            temp_cf_sheet.cell(row=cf_row, column=c).font = font_regular
        cf_row += 1
        
        # 2. OUTFLOW COGS row
        temp_cf_sheet.cell(row=cf_row, column=1, value=date_str)
        temp_cf_sheet.cell(row=cf_row, column=2, value=day_name)
        temp_cf_sheet.cell(row=cf_row, column=3, value="OUTFLOW")
        temp_cf_sheet.cell(row=cf_row, column=4, value="Daily COGS — Inventory Cost")
        temp_cf_sheet.cell(row=cf_row, column=5, value=f"COGS-{month_num:02d}-{day:02d}")
        temp_cf_sheet.cell(row=cf_row, column=6, value=f"=SUMPRODUCT('Sales Data'!{sales_col_letter}3:{sales_col_letter}163,'Sales Data'!$K$3:$K$163)")
        temp_cf_sheet.cell(row=cf_row, column=6).number_format = '#,##0.00'
        
        for c in range(1, 7):
            temp_cf_sheet.cell(row=cf_row, column=c).border = border_thin
            temp_cf_sheet.cell(row=cf_row, column=c).font = font_regular
        cf_row += 1
        
        # 3. OUTFLOW Expense row
        matching_expenses = [e for e in expense_data if e["day"] == day]
        for exp in matching_expenses:
            temp_cf_sheet.cell(row=cf_row, column=1, value=date_str)
            temp_cf_sheet.cell(row=cf_row, column=2, value=day_name)
            temp_cf_sheet.cell(row=cf_row, column=3, value="OUTFLOW")
            temp_cf_sheet.cell(row=cf_row, column=4, value=f"Operating Expense: {exp['desc']}")
            temp_cf_sheet.cell(row=cf_row, column=5, value=f"EXP-{month_num:02d}-{day:02d}")
            temp_cf_sheet.cell(row=cf_row, column=6, value=exp['amt'])
            temp_cf_sheet.cell(row=cf_row, column=6).number_format = '#,##0.00'
            
            for c in range(1, 7):
                temp_cf_sheet.cell(row=cf_row, column=c).border = border_thin
                temp_cf_sheet.cell(row=cf_row, column=c).font = font_regular
            cf_row += 1
            
    cf_row += 1
    temp_cf_sheet.cell(row=cf_row, column=1, value="CASH FLOW SUMMARY").font = font_bold
    cf_row += 1
    
    inflow_row = cf_row
    temp_cf_sheet.cell(row=cf_row, column=1, value="Total Cash Inflows (Sales Revenue)").font = font_bold
    temp_cf_sheet.cell(row=cf_row, column=6, value=f'=SUMIF(C3:C{inflow_row-2},"INFLOW",F3:F{inflow_row-2})').font = font_bold
    temp_cf_sheet.cell(row=cf_row, column=6).number_format = '#,##0.00'
    cf_row += 1
    
    outflow_row = cf_row
    temp_cf_sheet.cell(row=cf_row, column=1, value="Total Cash Outflows (COGS + Expenses)").font = font_bold
    temp_cf_sheet.cell(row=cf_row, column=6, value=f'=SUMIF(C3:C{outflow_row-1},"OUTFLOW",F3:F{outflow_row-1})').font = font_bold
    temp_cf_sheet.cell(row=cf_row, column=6).number_format = '#,##0.00'
    cf_row += 1
    
    net_cf_row = cf_row
    temp_cf_sheet.cell(row=cf_row, column=1, value="NET CASH FLOW").font = font_bold
    temp_cf_sheet.cell(row=cf_row, column=6, value=f'=F{inflow_row}-F{outflow_row}').font = font_bold
    temp_cf_sheet.cell(row=cf_row, column=6).number_format = '#,##0.00'
    
    # Format Summary rows
    for r in range(inflow_row, net_cf_row + 1):
        for c in range(1, 7):
            cell = temp_cf_sheet.cell(row=r, column=c)
            cell.fill = fill_summary
            if r == net_cf_row:
                cell.border = border_summary
                
    # -------------------------------------------------------------
    # 7. Update Financial Summary Sheet
    # -------------------------------------------------------------
    print("  Processing Financial Summary...")
    temp_fs_sheet = wb["Financial Summary"]
    temp_fs_sheet.cell(row=1, column=1).value = f"Al-Noor Trading — Monthly Financial Summary ({month_name} {year_val})"
    temp_fs_sheet.cell(row=2, column=1).value = f"Reporting Period: 01 {month_name} {year_val} – {num_days:02d} {month_name} {year_val}  |  Currency: PKR  |  Products: 161 Active SKUs"
    temp_fs_sheet.cell(row=25, column=2).value = f"='Revenue Analysis'!E164/{num_days}"
    
    # -------------------------------------------------------------
    # 8. Update KPI Dashboard Data Sheet
    # -------------------------------------------------------------
    print("  Processing KPI Dashboard Data...")
    temp_kpi_sheet = wb["KPI Dashboard Data"]
    temp_kpi_sheet.cell(row=1, column=1).value = f"Al-Noor Trading — KPI Dashboard Data ({month_name} {year_val})"
    temp_kpi_sheet.cell(row=19, column=1).value = f"Source: Sales Data sheet (161 SKUs, {num_days} days {month_name} {year_val}). Expenses: SUPPLIES only, PKR 31,589.99 (12 active). All vendors Pakistani (Falak Foods, Gul Ahmed, GSK Pakistan, Bazm-e-iqbal, Grey). Tax = PKR 0.00 on all records."
    
    # Save Workbook with error handling for locks
    out_path = PROJECT_ROOT / output_filename
    try:
        wb.save(out_path)
        print(f"  Saved: {output_filename}")
    except PermissionError:
        print(f"  Warning: Could not save '{output_filename}' because it is open in Excel or another program.")
    except Exception as e:
        print(f"  Error saving '{output_filename}': {e}")
        
    wb.close()
    in_wb.close()

def datetime_month_number(month_name):
    months = {
        "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
        "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12
    }
    return months[month_name]

# Run for all months
months_to_run = [
    ("January", "AlNoor_January_2026.xlsx", "AlNoor_Financial_Summary_January_2026.xlsx"),
    ("February", "AlNoor_February_2026.xlsx", "AlNoor_Financial_Summary_February_2026.xlsx"),
    ("March", "AlNoor_March_2026.xlsx", "AlNoor_Financial_Summary_March_2026.xlsx"),
    ("April", "AlNoor_April_2026.xlsx", "AlNoor_Financial_Summary_April_2026.xlsx"),
    ("June", "AlNoor_June_2026.xlsx", "AlNoor_Financial_Summary_June_2026.xlsx"),
]

for m_name, in_file, out_file in months_to_run:
    try:
        generate_monthly_summary(m_name, 2026, in_file, out_file)
    except Exception as e:
        print(f"Failed to generate summary for {m_name}: {e}")
    
print("\nAll financial summary files processed.")
