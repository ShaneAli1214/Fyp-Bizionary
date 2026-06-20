import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Daily Totals"]
print("Header Row 2:")
print([sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)])
print("Row 3 detail:")
for c in range(1, sheet.max_column + 1):
    val = sheet.cell(row=3, column=c).value
    print(f"Col {c}: value={val}, type={type(val)}")
wb.close()
