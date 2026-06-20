import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Category Summary"]
print("Row | Col A | Col B | Col C")
for r in range(1, sheet.max_row + 1):
    val_a = sheet.cell(row=r, column=1).value
    val_b = sheet.cell(row=r, column=2).value
    val_c = sheet.cell(row=r, column=3).value
    print(f"{r} | {val_a} | {val_b} | {val_c}")
wb.close()
