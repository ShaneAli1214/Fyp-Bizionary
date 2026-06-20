import openpyxl

wb = openpyxl.load_workbook("output/AlNoor_April_2026.xlsx")
sheet = wb["Category Summary"]
print("Header Row 2:")
print([sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)])
print("Row 3 detail:")
for c in range(1, sheet.max_column + 1):
    print(f"Col {c}: value={sheet.cell(row=3, column=c).value}, type={type(sheet.cell(row=3, column=c).value)}")
print("GRAND TOTAL row detail:")
r = sheet.max_row
for c in range(1, sheet.max_column + 1):
    print(f"Col {c}: value={sheet.cell(row=r, column=c).value}, type={type(sheet.cell(row=r, column=c).value)}")
wb.close()
