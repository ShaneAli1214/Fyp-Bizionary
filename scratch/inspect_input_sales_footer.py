import openpyxl

wb = openpyxl.load_workbook("output/AlNoor_April_2026.xlsx")
sheet = wb["Sales Data"]
print("Max row:", sheet.max_row)
for r in range(sheet.max_row - 5, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")
wb.close()
