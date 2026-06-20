import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary.xlsx")
sheet = wb["Revenue Analysis"]
r = 217
print(f"Row {r} detail:")
for c in range(1, sheet.max_column + 1):
    print(f"Col {c}: value={sheet.cell(row=r, column=c).value}, type={type(sheet.cell(row=r, column=c).value)}")
wb.close()
