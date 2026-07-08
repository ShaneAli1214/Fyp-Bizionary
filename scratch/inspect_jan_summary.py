import openpyxl

wb = openpyxl.load_workbook("AlNoor_Financial_Summary_January_2026.xlsx", data_only=True)
if 'Expense Tracker' in wb.sheetnames:
    sheet = wb['Expense Tracker']
    print("Expense Tracker columns:", [sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)])
    for r in range(3, 10):
        print(f"Row {r}: {[sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]}")
wb.close()
